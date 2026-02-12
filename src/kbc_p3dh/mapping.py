"""
Parse EBA Annotated Table Layout Excel files to build a datapoint mapping.

Each Excel sheet represents one EBA Pillar 3 template.  The layout is:
  Row 1   : Template title
  Row 4   : "Columns" header + dimension properties
  Row 5   : Human-readable column labels  (a. T, b. T-1 …)
  Row 6   : Column codes                  (0010, 0020 …)
  Row 7+  : Data rows  (col A = "Rows" for first, col B = row label,
            col C = row code, col D+ = datapoint cells)

Datapoint cells look like  "3530270_x000D_\\n€£$"  where:
  - The number before "_x000D_" is the datapoint ID  (→ dp3530270)
  - €£$ = monetary,  % = percentage,  # = count,  yyyy-mm-dd = date
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ── Paths ──────────────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parents[2]          # project root
MAPPING_DIR = _ROOT / "Mapping"
DATA_DIR = _ROOT / "P3DH downloads"
CODI_XLSX = MAPPING_DIR / "20250519 Annotated Table Layout  CODISPILLAR3 4.1.xlsx"
IRRBB_XLSX = MAPPING_DIR / "20250519 Annotated Table Layout  IRRBBDISPILLAR3 4.1.xlsx"


# ── helpers ────────────────────────────────────────────────────────────────────
_DP_RE = re.compile(r"^(\d+)")  # leading digits = datapoint id

UNIT_MAP = {
    "€£$": "monetary",
    "%": "percentage",
    "#": "count",
    "yyyy-mm-dd": "date",
}


def _parse_dp_cell(cell_value: Any) -> tuple[str | None, str]:
    """Return (dp_id, unit) from a cell value like '3530270_x000D_\\n€£$'."""
    if cell_value is None:
        return None, ""
    text = str(cell_value).strip()
    m = _DP_RE.match(text)
    if not m:
        return None, ""
    dp_id = f"dp{m.group(1)}"
    unit = ""
    for marker, uname in UNIT_MAP.items():
        if marker in text:
            unit = uname
            break
    return dp_id, unit


# ── TOC parser ─────────────────────────────────────────────────────────────────
def parse_toc(xlsx_path: Path) -> dict[str, str]:
    """Return {sheet_code: template_title} from the TOC sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["TOC"]
    toc: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        code = row[1] if len(row) > 1 else None
        title = row[2] if len(row) > 2 else None
        if code and title:
            toc[str(code).strip()] = str(title).strip()
    wb.close()
    return toc


# ── Single-sheet parser ───────────────────────────────────────────────────────
def _parse_sheet(ws, sheet_name: str, toc: dict[str, str]) -> list[dict]:
    """Parse one annotated table layout sheet → list of datapoint records."""
    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 7:
        return []

    # Template name from row 1 or TOC
    template_title = toc.get(sheet_name, str(rows_raw[0][0] or sheet_name))

    # Column labels (row index 4, 0-based) and codes (row index 5)
    col_labels_row = rows_raw[4] if len(rows_raw) > 4 else ()
    col_codes_row = rows_raw[5] if len(rows_raw) > 5 else ()

    # Build column info starting from col index 3 (col D onwards)
    col_info: list[tuple[str, str]] = []  # (label, code)
    for ci in range(3, max(len(col_labels_row), len(col_codes_row))):
        label = str(col_labels_row[ci]).strip() if ci < len(col_labels_row) and col_labels_row[ci] else ""
        code = str(col_codes_row[ci]).strip() if ci < len(col_codes_row) and col_codes_row[ci] else ""
        col_info.append((label, code))

    # Data rows: start from row index 6 (row 7), stop at "Main Property"
    records: list[dict] = []
    for ri in range(6, len(rows_raw)):
        row = rows_raw[ri]
        if not row or len(row) < 4:
            continue
        # Stop when we hit metadata rows
        row_label = str(row[1]).strip() if row[1] else ""
        if row_label in ("Main Property", "") and row[0] is None:
            # Could be a metadata row or section header – skip
            if row_label == "Main Property":
                break
            # Section header like "0035 Risk-weighted exposure amounts"
            if row[2] is None:
                continue
        row_code = str(row[2]).strip() if row[2] else ""
        if not row_code:
            continue  # section header, skip

        # Extract datapoints from col D onwards
        for ci in range(3, len(row)):
            dp_id, unit = _parse_dp_cell(row[ci])
            if dp_id is None:
                continue
            col_idx = ci - 3
            col_label, col_code = col_info[col_idx] if col_idx < len(col_info) else ("", "")
            records.append({
                "datapoint": dp_id,
                "template": sheet_name,
                "template_title": template_title,
                "row_label": row_label,
                "row_code": row_code,
                "col_label": col_label,
                "col_code": col_code,
                "unit": unit,
            })
    return records


# ── Full mapping builder ──────────────────────────────────────────────────────
def build_mapping() -> pd.DataFrame:
    """Parse both CODI and IRRBB Excel files → unified mapping DataFrame.

    Columns: datapoint, template, template_title, row_label, row_code,
             col_label, col_code, unit
    """
    all_records: list[dict] = []

    for xlsx_path in (CODI_XLSX, IRRBB_XLSX):
        if not xlsx_path.exists():
            continue
        toc = parse_toc(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            if sname == "TOC":
                continue
            ws = wb[sname]
            records = _parse_sheet(ws, sname, toc)
            all_records.extend(records)
        wb.close()

    df = pd.DataFrame(all_records)
    # Remove duplicates (same dp can appear in parametrised sheets)
    df = df.drop_duplicates(subset=["datapoint", "template"])
    return df
