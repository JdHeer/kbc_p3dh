"""
Universal P3DH data ingestion — build a single SQLite database of fact-level data
from any number of P3DH download folders, mapped via EBA Annotated Table Layouts.

Each row in the `facts` table represents **one reported value** with full metadata:
entity, period, template, row, column, value — ready for cross-entity queries.

Usage
─────
    # Ingest one folder (auto-detects entity, period, module):
    uv run python ingest_p3dh.py  "P3DH downloads"

    # Ingest multiple folders at once:
    uv run python ingest_p3dh.py  "downloads/ABN_2025H1"  "downloads/KBC_2025H1"

    # Custom DB path:
    uv run python ingest_p3dh.py  --db mydata.db  "P3DH downloads"

The script is idempotent: re-running with the same folder replaces that
entity+period combination (upsert by delete-then-insert).
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

# ── Constants ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
MAPPING_DIR = ROOT / "Mapping"
DB_DEFAULT = ROOT / "p3dh.db"

# Well-known LEI → short entity name
LEI_NAMES: dict[str, str] = {
    "BFXS5XCH7N0Y05NIXW11": "ABN AMRO",
    "6V5X0Z0Y1OCLKZXOI394": "ABN AMRO (solo)",
    "549300DIFDNI4Q13  ":    "KBC Group",         # placeholder, update as needed
    "213800MBWEIJDM5CU638": "KBC Group",
    "LSGM84136ACA92XCN876": "ING Group",
    "JLS56RAMYQZECFUF2G44": "Rabobank",
    "DG3RU1DBUFHT4ZF9WN62": "BNP Paribas",
    "R0MUWSFPU8MPRO8K5P83": "HSBC",
    "529900HNOAA1KXQJUQ27": "Deutsche Bank",
    "O2RNE8IBXP4R0TD8PU41": "Goldman Sachs",
    "5493006MHB84DD0ZWV18": "Barclays",
    "B4TYDEB6GKMZO031MB27": "UBS",
}

# Module keyword → mapping Excel file (filename substring match)
MODULE_MAPPING_FILES: dict[str, str] = {
    "findis":  "FINDISPILLAR3",
    "codi":    "CODISPILLAR3",
    "irrbb":   "IRRBBDISPILLAR3",
    "esg":     "ESGDISPILLAR3",
    "gsii":    "GSIIDISPILLAR3",
    "mrel":    "MRELTLACDISPILLAR3",
    "rem":     "REMDISPILLAR3",
    "p3dh":    "P3DHPILLAR3",
}


# ── Datapoint cell parser ─────────────────────────────────────────────────────
_DP_RE = re.compile(r"^(\d+)")

UNIT_MAP = {"€£$": "monetary", "%": "percentage", "#": "count", "yyyy-mm-dd": "date"}


def _parse_dp_cell(val: Any) -> tuple[str | None, str]:
    """Extract (dp_id, unit) from a cell like '3530270_x000D_\\n€£$'."""
    if val is None:
        return None, ""
    text = str(val).strip()
    m = _DP_RE.match(text)
    if not m:
        return None, ""
    dp_id = f"dp{m.group(1)}"
    unit = next((u for marker, u in UNIT_MAP.items() if marker in text), "")
    return dp_id, unit


# ── Mapping parser ────────────────────────────────────────────────────────────
def _find_mapping_xlsx(module: str) -> Path | None:
    """Find the mapping Excel for a given module keyword."""
    key = module.lower().strip("/")
    for mod_key, file_substr in MODULE_MAPPING_FILES.items():
        if mod_key in key:
            for p in MAPPING_DIR.glob("*.xlsx"):
                if file_substr in p.name:
                    return p
    return None


def _parse_toc(xlsx_path: str | Path) -> dict[str, str]:
    """Return {sheet_code: template_title} from the TOC sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["TOC"]
    toc: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        code = row[1] if len(row) > 1 else None
        title = row[2] if len(row) > 2 else None
        if code and title:
            toc[str(code).strip()] = str(title).strip()
    wb.close()
    return toc


def _parse_mapping_sheet(ws, sheet_name: str, toc: dict[str, str]) -> list[dict]:
    """Parse one annotated table layout sheet → list of datapoint dicts.

    Captures multi-level column headers by forward-filling parent labels.
    Dynamically detects header rows by looking for the column-code row
    (the row containing 4-digit codes like "0010", "0020") and the first
    data row (starts with "Rows" in col A or has a row code in col C).
    """
    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 5:
        return []

    template_title = toc.get(sheet_name, str(rows_raw[0][0] or sheet_name))

    # ── Find row indices dynamically ────────────────────────────────────────
    # Scan for the column-code row: a row where col D+ contain 4-digit strings
    code_row_idx = None
    for ri in range(3, min(len(rows_raw), 12)):
        row = rows_raw[ri]
        if row and len(row) > 3:
            # Check if col D (index 3) looks like a 4-digit code
            v = str(row[3]).strip() if row[3] else ""
            if re.match(r"^\d{4}$", v):
                code_row_idx = ri
                break

    if code_row_idx is None:
        # Fallback: assume standard layout (codes at row index 7)
        code_row_idx = 7 if len(rows_raw) > 7 else len(rows_raw) - 1

    # Header rows are the 1-3 rows above the code row
    # Typically: header_row = code_row_idx - 3, group_row = code_row_idx - 2, sub_row = code_row_idx - 1
    header_row = rows_raw[code_row_idx - 3] if code_row_idx >= 3 and code_row_idx - 3 >= 0 else ()
    group_row  = rows_raw[code_row_idx - 2] if code_row_idx >= 2 else ()
    sub_row    = rows_raw[code_row_idx - 1] if code_row_idx >= 1 else ()
    code_row   = rows_raw[code_row_idx]

    # Data rows start right after the code row
    data_start = code_row_idx + 1

    # ── Column headers ──────────────────────────────────────────────────────
    max_ci = max(len(header_row), len(group_row), len(sub_row), len(code_row))
    col_info: list[dict[str, str] | None] = []

    last_header = ""
    last_group = ""
    for ci in range(3, max_ci):
        h = str(header_row[ci]).strip() if ci < len(header_row) and header_row[ci] else ""
        g = str(group_row[ci]).strip()  if ci < len(group_row) and group_row[ci]  else ""
        s = str(sub_row[ci]).strip()    if ci < len(sub_row) and sub_row[ci]    else ""
        c = str(code_row[ci]).strip()   if ci < len(code_row) and code_row[ci]   else ""

        # Skip dimension property columns (start with "(q")
        if g.startswith("(q") or h.startswith("(q"):
            col_info.append(None)
            continue

        if h:
            last_header = h
        if g:
            last_group = g

        col_label = s or g or h or ""
        col_parent = last_group if s else (last_header if g else "")

        col_info.append({
            "col_code": c,
            "col_header": last_header,
            "col_group": last_group if g or s else "",
            "col_sub": s,
            "col_label": col_label,
            "col_parent": col_parent,
        })

    # ── Data rows ───────────────────────────────────────────────────────────
    records: list[dict] = []

    for ri in range(data_start, len(rows_raw)):
        row = rows_raw[ri]
        if not row or len(row) < 4:
            continue
        row_label = str(row[1]).strip() if row[1] else ""
        if row_label == "Main Property":
            break
        row_code = str(row[2]).strip() if row[2] else ""
        if not row_code:
            continue

        for ci in range(3, len(row)):
            dp_id, unit = _parse_dp_cell(row[ci])
            if dp_id is None:
                continue
            col_idx = ci - 3
            if col_idx >= len(col_info) or col_info[col_idx] is None:
                continue
            ci_dict = col_info[col_idx]
            records.append({
                "datapoint":      dp_id,
                "template":       sheet_name,
                "template_title": template_title,
                "row_label":      row_label,
                "row_code":       row_code,
                "col_code":       ci_dict["col_code"],
                "col_label":      ci_dict["col_label"],
                "col_parent":     ci_dict["col_parent"],
                "col_header":     ci_dict["col_header"],
                "col_group":      ci_dict["col_group"],
                "col_sub":        ci_dict["col_sub"],
                "unit":           unit,
            })
    return records


def build_mapping(xlsx_path: str | Path) -> dict[str, dict]:
    """Parse a mapping Excel → {datapoint: metadata_dict}.

    Returns a dict keyed by datapoint ID for fast lookup during merge.
    Also returns the records list for bulk insertion.
    """
    import openpyxl

    toc = _parse_toc(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_records: list[dict] = []
    for sname in wb.sheetnames:
        if sname == "TOC":
            continue
        ws = wb[sname]
        records = _parse_mapping_sheet(ws, sname, toc)
        all_records.extend(records)
    wb.close()

    # Deduplicate by (datapoint, template)
    seen = set()
    unique: list[dict] = []
    for r in all_records:
        key = (r["datapoint"], r["template"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


# ── Metadata reader ───────────────────────────────────────────────────────────
def _read_parameters(folder: Path) -> dict[str, str]:
    """Read parameters.csv → dict."""
    pfile = folder / "parameters.csv"
    if not pfile.exists():
        return {}
    params = {}
    with open(pfile, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            params[row["name"]] = row["value"]
    return params


def _read_module(folder: Path) -> str:
    """Extract module name from report.json."""
    rfile = folder / "report.json"
    if not rfile.exists():
        return ""
    data = json.loads(rfile.read_text(encoding="utf-8"))
    extends = data.get("documentInfo", {}).get("extends", [])
    # Extract module from URL like ".../mod/findis.json"
    for url in extends:
        m = re.search(r"/mod/(\w+)\.json", url)
        if m:
            return m.group(1)
    return ""


def _lei_to_name(lei: str) -> str:
    """Resolve LEI to a short entity name."""
    # Strip prefixes like "rs:", "lei:" and suffixes like ".CON"
    clean = re.sub(r"^(rs:|lei:)", "", lei, flags=re.IGNORECASE)
    clean = clean.split(".")[0].strip()
    return LEI_NAMES.get(clean, clean)


# ── CSV data loader ───────────────────────────────────────────────────────────
def _load_csv_data(folder: Path) -> list[tuple[str, str, str]]:
    """Read all k_*.csv files → list of (datapoint, factValue, file_stem)."""
    facts = []
    for csv_path in sorted(folder.glob("k_*.csv")):
        stem = csv_path.stem
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                facts.append((row["datapoint"], row["factValue"], stem))
    return facts


# ── Database setup ────────────────────────────────────────────────────────────
CREATE_FACTS = """
CREATE TABLE IF NOT EXISTS facts (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    entity        TEXT NOT NULL,
    lei           TEXT NOT NULL,
    ref_period    TEXT NOT NULL,
    module        TEXT NOT NULL,
    currency      TEXT,
    template      TEXT NOT NULL,
    template_title TEXT,
    row_code      TEXT,
    row_label     TEXT,
    col_code      TEXT,
    col_label     TEXT,
    col_parent    TEXT,
    col_header    TEXT,
    col_group     TEXT,
    col_sub       TEXT,
    file          TEXT,
    datapoint     TEXT NOT NULL,
    value         REAL,
    value_raw     TEXT,
    value_eur_m   REAL,
    unit          TEXT
);
"""

CREATE_INDEXES = [
    "CREATE INDEX IF NOT EXISTS idx_facts_entity     ON facts(entity);",
    "CREATE INDEX IF NOT EXISTS idx_facts_period     ON facts(ref_period);",
    "CREATE INDEX IF NOT EXISTS idx_facts_template   ON facts(template);",
    "CREATE INDEX IF NOT EXISTS idx_facts_datapoint  ON facts(datapoint);",
    "CREATE INDEX IF NOT EXISTS idx_facts_entity_period ON facts(entity, ref_period);",
    "CREATE INDEX IF NOT EXISTS idx_facts_module     ON facts(module);",
]

INSERT_FACT = """
INSERT INTO facts (
    entity, lei, ref_period, module, currency,
    template, template_title, row_code, row_label,
    col_code, col_label, col_parent, col_header, col_group, col_sub,
    file, datapoint, value, value_raw, value_eur_m, unit
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


def _init_db(db_path: Path) -> sqlite3.Connection:
    """Create / open the database and ensure schema exists."""
    con = sqlite3.connect(str(db_path))
    con.execute("PRAGMA journal_mode=WAL")
    con.execute(CREATE_FACTS)
    for idx_sql in CREATE_INDEXES:
        con.execute(idx_sql)
    con.commit()
    return con


# ── Main ingestion ────────────────────────────────────────────────────────────
def ingest_folder(folder: Path, db_path: Path = DB_DEFAULT) -> None:
    """Ingest one P3DH download folder into the facts database."""
    folder = folder.resolve()
    print(f"\n{'─' * 70}")
    print(f"Ingesting: {folder}")

    # 1. Read metadata
    params = _read_parameters(folder)
    module = _read_module(folder)
    lei_raw = params.get("entityID", "UNKNOWN")
    # Strip prefixes like "rs:", "lei:" and suffixes like ".CON"
    lei = re.sub(r"^(rs:|lei:)", "", lei_raw, flags=re.IGNORECASE)
    lei = lei.split(".")[0].strip()
    entity = _lei_to_name(lei_raw)
    ref_period = params.get("refPeriod", "UNKNOWN")
    currency = params.get("baseCurrency", "EUR")

    print(f"  Entity:     {entity} ({lei})")
    print(f"  Period:     {ref_period}")
    print(f"  Module:     {module}")
    print(f"  Currency:   {currency}")

    # 2. Find & parse mapping
    mapping_path = _find_mapping_xlsx(module)
    if mapping_path is None:
        print(f"  ERROR: No mapping Excel found for module '{module}'")
        return
    print(f"  Mapping:    {mapping_path.name}")

    mapping_records = build_mapping(mapping_path)
    # Build lookup: datapoint → record
    dp_lookup: dict[str, dict] = {}
    for rec in mapping_records:
        dp_lookup.setdefault(rec["datapoint"], rec)
    print(f"  Datapoints: {len(dp_lookup):,} definitions from mapping")

    # 3. Load CSV data
    csv_facts = _load_csv_data(folder)
    print(f"  CSV facts:  {len(csv_facts):,} raw values")

    # 4. Match & prepare rows
    matched = 0
    unmatched = 0
    rows_to_insert: list[tuple] = []

    for dp, raw_val, file_stem in csv_facts:
        meta = dp_lookup.get(dp)
        if meta is None:
            unmatched += 1
            continue
        matched += 1
        try:
            numeric = float(raw_val)
        except (ValueError, TypeError):
            numeric = None
        value_eur_m = numeric / 1e6 if numeric is not None else None
        rows_to_insert.append((
            entity, lei, ref_period, module, currency,
            meta["template"], meta["template_title"],
            meta["row_code"], meta["row_label"],
            meta["col_code"], meta["col_label"], meta["col_parent"],
            meta["col_header"], meta["col_group"], meta["col_sub"],
            file_stem, dp, numeric, raw_val, value_eur_m, meta["unit"],
        ))

    print(f"  Matched:    {matched:,} / Unmatched: {unmatched}")

    # 5. Write to DB (idempotent: delete existing entity+period+module first)
    con = _init_db(db_path)
    con.execute(
        "DELETE FROM facts WHERE lei = ? AND ref_period = ? AND module = ?",
        (lei, ref_period, module),
    )
    con.executemany(INSERT_FACT, rows_to_insert)
    con.commit()

    total = con.execute("SELECT COUNT(*) FROM facts").fetchone()[0]
    entities = con.execute("SELECT COUNT(DISTINCT entity) FROM facts").fetchone()[0]
    periods = con.execute("SELECT COUNT(DISTINCT ref_period) FROM facts").fetchone()[0]
    con.close()

    print(f"  Inserted:   {len(rows_to_insert):,} facts")
    print(f"  DB totals:  {total:,} facts across {entities} entities, {periods} periods")


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Ingest P3DH download folders into a universal SQLite database"
    )
    parser.add_argument(
        "folders", nargs="+", type=Path,
        help="One or more P3DH download folders (each containing k_*.csv + parameters.csv + report.json)"
    )
    parser.add_argument(
        "--db", type=Path, default=DB_DEFAULT,
        help=f"Output database path (default: {DB_DEFAULT.name})"
    )
    args = parser.parse_args()

    print(f"Database: {args.db}")
    for folder in args.folders:
        if not folder.is_dir():
            print(f"WARNING: {folder} is not a directory, skipping")
            continue
        ingest_folder(folder, args.db)
    print(f"\nDone.")


if __name__ == "__main__":
    main()
